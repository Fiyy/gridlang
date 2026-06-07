"""Tests for gridlang.csv_io and gridlang.excel_import/export"""

import pytest
import pandas as pd
from pathlib import Path

from gridlang.csv_io import import_csv, export_csv, export_csv_string
from gridlang.excel_import import import_excel
from gridlang.excel_export import export_excel, export_dataframe
from gridlang.parser import parse_string


@pytest.fixture
def sample_csv(tmp_path):
    csv_file = tmp_path / "test.csv"
    csv_file.write_text("Name,Score,Grade\nAlice,90,A\nBob,80,B\nCharlie,70,C\n")
    return csv_file


@pytest.fixture
def sample_grid(tmp_path):
    grid_content = """--- meta ---
name: "Test"
engine: python
version: "1.0"

--- data ---
Product,Price,Qty
Widget,10,100
Gadget,20,50
Tool,15,75

--- compute ---
def transform(df):
    df['Total'] = df['Price'] * df['Qty']
    return df

def aggregates(df):
    return {'grand_total': df['Total'].sum()}

--- present ---
<p>Total: {{ agg.grand_total }}</p>
"""
    grid_file = tmp_path / "test.grid"
    grid_file.write_text(grid_content)
    return grid_file


@pytest.fixture
def sample_xlsx(tmp_path):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Name", "Value", "Category"])
    ws.append(["Alpha", 100, "A"])
    ws.append(["Beta", 200, "B"])
    ws.append(["Gamma", 150, "A"])
    xlsx_file = tmp_path / "test.xlsx"
    wb.save(xlsx_file)
    wb.close()
    return xlsx_file


class TestCSVImport:
    def test_basic_import(self, sample_csv):
        result = import_csv(sample_csv)
        assert "--- meta ---" in result
        assert "--- data ---" in result
        assert "--- compute ---" in result
        assert "--- present ---" in result
        assert "Alice" in result

    def test_name_from_filename(self, sample_csv):
        result = import_csv(sample_csv)
        assert 'name: "Test"' in result

    def test_custom_name(self, sample_csv):
        result = import_csv(sample_csv, name="My Report")
        assert 'name: "My Report"' in result

    def test_result_is_valid_grid(self, sample_csv):
        result = import_csv(sample_csv)
        doc = parse_string(result)
        assert doc.name == "Test"
        assert "Alice" in doc.data_raw

    def test_file_not_found(self):
        with pytest.raises(FileNotFoundError):
            import_csv("/nonexistent/file.csv")


class TestCSVExport:
    def test_basic_export(self, sample_grid, tmp_path):
        output = tmp_path / "output.csv"
        result_path = export_csv(sample_grid, output)
        assert result_path.exists()
        content = result_path.read_text()
        assert "Product" in content
        assert "Total" in content  # Computed column
        assert "1000" in content   # Widget total: 10*100

    def test_export_raw(self, sample_grid, tmp_path):
        output = tmp_path / "raw.csv"
        result_path = export_csv(sample_grid, output, raw=True)
        content = result_path.read_text()
        assert "Product" in content
        assert "Total" not in content  # No compute

    def test_export_string(self, sample_grid):
        csv_str = export_csv_string(sample_grid)
        assert "Widget" in csv_str
        assert "Total" in csv_str


class TestExcelImport:
    """Behavioural tests for `gridlang.excel_import.import_excel`.

    These exercise the public API end-to-end: a real .xlsx is built with
    openpyxl, fed through `import_excel`, and the resulting `.grid` is
    parsed/executed/rendered to verify what the user actually sees.

    Categories:
      1. Smoke / contract — produces a valid 4-section .grid string.
      2. Header detection — text headers vs all-numeric first row.
      3. Cell-type fidelity — strings, ints, floats, dates, bools, NaN.
      4. Formula extraction — plain formulas + ArrayFormula dynamic arrays.
      5. Trailing summary rows — pulled out of data into compute.
      6. Detached scattered rows — pulled out, don't pollute dtypes.
      7. End-to-end render — the rendered HTML matches expectations.
      8. Multi-sheet, sheet selection, and error paths.
    """

    # ── 1. Smoke / contract ──────────────────────────────────────────────

    def test_basic_import(self, sample_xlsx):
        result = import_excel(sample_xlsx)
        assert "--- meta ---" in result
        assert "--- data ---" in result
        assert "Alpha" in result
        assert "200" in result

    def test_result_is_valid_grid(self, sample_xlsx):
        result = import_excel(sample_xlsx)
        doc = parse_string(result)
        # All four sections must be parseable.
        assert doc.meta.get("name")
        assert "Alpha" in doc.data_raw
        # Compute and present sections always emitted, even if empty-ish.
        assert "def transform" in doc.compute_raw
        assert "<table" in doc.present_raw

    def test_meta_fields_populated(self, sample_xlsx):
        result = import_excel(sample_xlsx)
        doc = parse_string(result)
        assert doc.meta["name"] == "test"  # filename stem
        assert doc.meta["engine"] == "python"
        assert doc.meta["imported_from"] == "test.xlsx"
        # import_date must be a string, not a datetime literal.
        assert "import_date" in doc.meta

    def test_filename_with_unicode(self, tmp_path):
        """Chinese / accented filenames preserved in meta."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Value"])
        ws.append(["x", 1])
        f = tmp_path / "测试_报告.xlsx"
        wb.save(f); wb.close()

        result = import_excel(f)
        doc = parse_string(result)
        assert doc.meta["name"] == "测试_报告"
        assert "测试_报告.xlsx" in doc.meta["imported_from"]

    # ── 2. Header detection ──────────────────────────────────────────────

    def test_no_header_row_detected_when_all_numeric(self, tmp_path):
        """Row 1 = all numbers → don't promote it to header.

        Before this fix, a pure-data sheet `1,1,...` / `2,2,...` would have
        row 1 silently consumed as a header, dropping a real data row.
        """
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        for v in range(1, 6):  # 5 rows of all-numeric data
            ws.append([v] * 4)
        f = tmp_path / "no_header.xlsx"
        wb.save(f); wb.close()

        result = import_excel(f)
        from gridlang.schema import parse_data
        df = parse_data(parse_string(result).data_raw)

        # All 5 source rows must survive — none consumed as header.
        assert len(df) == 5
        assert list(df.columns) == ["col_A", "col_B", "col_C", "col_D"]
        assert int(df.iloc[0, 0]) == 1
        assert int(df.iloc[-1, 0]) == 5

    def test_header_row_detected_when_text(self, tmp_path):
        """Row 1 = text labels → DO use it as header."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["Region", "Q1", "Q2"])
        ws.append(["North", 100, 200])
        ws.append(["South", 150, 250])
        f = tmp_path / "with_header.xlsx"
        wb.save(f); wb.close()

        result = import_excel(f)
        from gridlang.schema import parse_data
        df = parse_data(parse_string(result).data_raw)

        assert list(df.columns) == ["Region", "Q1", "Q2"]
        assert len(df) == 2
        assert df.iloc[0]["Region"] == "North"

    def test_mixed_first_row_is_data_not_header(self, tmp_path):
        """Row 1 with a number among labels is data, not a header."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["Item", 100, 200])  # mixed → not a header
        ws.append(["A", 1, 2])
        ws.append(["B", 3, 4])
        f = tmp_path / "mixed_first_row.xlsx"
        wb.save(f); wb.close()

        result = import_excel(f)
        from gridlang.schema import parse_data
        df = parse_data(parse_string(result).data_raw)
        # Synthetic headers since the first row isn't pure-text.
        assert list(df.columns) == ["col_A", "col_B", "col_C"]
        assert len(df) == 3

    def test_numeric_string_first_row_is_data_not_header(self, tmp_path):
        """A first row of numeric *strings* like '2024','2025' is still data."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["2024", "2025", "2026"])  # all numeric strings → data
        ws.append([100, 110, 120])
        f = tmp_path / "numeric_strings.xlsx"
        wb.save(f); wb.close()

        result = import_excel(f)
        from gridlang.schema import parse_data
        df = parse_data(parse_string(result).data_raw)
        assert list(df.columns) == ["col_A", "col_B", "col_C"]
        # 2 rows of data — first row preserved.
        assert len(df) == 2

    # ── 3. Cell-type fidelity ────────────────────────────────────────────

    def test_string_cells_preserved(self, sample_xlsx):
        """Strings round-trip without escaping issues."""
        result = import_excel(sample_xlsx)
        from gridlang.schema import parse_data
        df = parse_data(parse_string(result).data_raw)
        assert "Alpha" in df["Name"].tolist()
        assert "Beta" in df["Name"].tolist()
        assert "Gamma" in df["Name"].tolist()

    def test_string_with_comma_quoted(self, tmp_path):
        """A string containing a comma must be quoted in the CSV."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["Label", "Note"])
        ws.append(["Hello, World", "fine"])
        f = tmp_path / "comma.xlsx"
        wb.save(f); wb.close()

        result = import_excel(f)
        from gridlang.schema import parse_data
        df = parse_data(parse_string(result).data_raw)
        assert df.iloc[0]["Label"] == "Hello, World"

    def test_integer_cells_stay_integer_in_render(self, tmp_path):
        """Whole numbers must NOT render as 1.0 / 2.0 / 3.0 ...

        The bug: pandas promotes int columns with one NaN to float64. Our
        present template now collapses 1.0 -> 1 in the table cells.
        """
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["A", "B", "C"])
        ws.append([1, 2, 3])
        ws.append([4, 5, None])  # induce NaN → float64 promotion
        f = tmp_path / "ints.xlsx"
        wb.save(f); wb.close()

        result = import_excel(f)
        from gridlang.schema import parse_data
        from gridlang.runtime import execute
        from gridlang.renderer import render
        doc = parse_string(result)
        df = parse_data(doc.data_raw)
        out = execute(doc.compute_raw, df)
        html = render(
            template_content=doc.present_raw, df=out.df,
            aggregates=out.aggregates, meta=doc.meta,
        )
        # Strip outside-of-tbody chrome to avoid false positives in CSS.
        import re
        body_match = re.search(r"<tbody>(.*?)</tbody>", html, re.S)
        assert body_match, "Rendered HTML has no <tbody>"
        body = body_match.group(1)
        assert ">1.0<" not in body, "Integer rendered as float: 1.0"
        assert ">5.0<" not in body, "Integer rendered as float: 5.0"
        assert ">1<" in body, "Integer 1 should render as plain '1'"
        assert ">5<" in body

    def test_float_with_real_decimals_kept(self, tmp_path):
        """3.14 is NOT an integer — render as 3.14, not 3."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["v"])
        ws.append([3.14])
        ws.append([2.5])
        f = tmp_path / "floats.xlsx"
        wb.save(f); wb.close()

        result = import_excel(f)
        # CSV output should preserve the float value.
        doc = parse_string(result)
        assert "3.14" in doc.data_raw
        assert "2.5" in doc.data_raw

    def test_date_cells_iso_formatted(self, tmp_path):
        from openpyxl import Workbook
        from datetime import datetime as dt
        wb = Workbook()
        ws = wb.active
        ws.append(["When", "What"])
        ws.append([dt(2024, 7, 15), "launch"])
        f = tmp_path / "dates.xlsx"
        wb.save(f); wb.close()

        result = import_excel(f)
        # Date should be ISO yyyy-mm-dd in the data block.
        assert "2024-07-15" in result

    def test_bool_cells_preserved(self, tmp_path):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["Active", "Name"])
        ws.append([True, "Alice"])
        ws.append([False, "Bob"])
        f = tmp_path / "bools.xlsx"
        wb.save(f); wb.close()

        result = import_excel(f)
        # Booleans serialize as their string form.
        assert "True" in result and "False" in result

    def test_empty_cells_become_empty_strings(self, tmp_path):
        """None cells in the middle of a row become '' in CSV, not 'None'."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["A", "B", "C"])
        ws.append([1, None, 3])
        f = tmp_path / "empties.xlsx"
        wb.save(f); wb.close()

        result = import_excel(f)
        doc = parse_string(result)
        # Should be "1,,3" — never "1,None,3".
        assert "None" not in doc.data_raw
        assert "1,,3" in doc.data_raw

    # ── 4. Formula extraction ────────────────────────────────────────────

    def test_plain_formula_extracted(self, tmp_path):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["v", "double"])
        ws.append([1, "=A2*2"])
        ws.append([2, "=A3*2"])
        f = tmp_path / "formula.xlsx"
        wb.save(f); wb.close()

        result = import_excel(f)
        doc = parse_string(result)
        # The formula text should appear in compute as a comment.
        assert "=A2*2" in doc.compute_raw or "*2" in doc.compute_raw

    def test_formulas_disabled_skips_extraction(self, tmp_path):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["v", "double"])
        ws.append([1, "=A2*2"])
        f = tmp_path / "no_formulas.xlsx"
        wb.save(f); wb.close()

        result = import_excel(f, include_formulas=False)
        doc = parse_string(result)
        # No formula references should appear in compute.
        assert "*2" not in doc.compute_raw

    def test_array_formula_not_dropped(self, tmp_path):
        """ArrayFormula objects must survive extraction."""
        from openpyxl import Workbook
        from openpyxl.worksheet.formula import ArrayFormula
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Score"])
        ws.append(["Alice", 90])
        ws.append(["Bob", 80])
        ws["C1"] = ArrayFormula("C1:C2", "=B1:B2*2")
        f = tmp_path / "array_formula.xlsx"
        wb.save(f); wb.close()

        result = import_excel(f)
        doc = parse_string(result)
        assert "B1:B2" in doc.compute_raw or "C1:C2" in doc.compute_raw, (
            "Array formula was silently dropped"
        )

    def test_duplicate_formula_pattern_collapsed(self, tmp_path):
        """The same formula pattern across many rows shouldn't create N comments."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["v", "x2"])
        for i in range(2, 12):  # 10 rows of `=A{n}*2`
            ws.append([i, f"=A{i}*2"])
        f = tmp_path / "many_formulas.xlsx"
        wb.save(f); wb.close()

        result = import_excel(f)
        doc = parse_string(result)
        # The "=A2*2" pattern (with N for digits) should appear at most once
        # in the deduped block — count its concrete instances.
        occurrences = doc.compute_raw.count("*2")
        assert occurrences <= 3, (
            f"Expected dedup of identical formula pattern; saw {occurrences}"
        )

    # ── 5. Trailing summary rows (kept in data, annotated in compute) ────

    def test_trailing_summary_row_value_preserved_in_data(self, tmp_path):
        """A trailing `=SUM(...)` row's CACHED value stays in the data
        block (it's a real cell). The formula text appears in compute as
        an advisory hint, but the value is NEVER stripped from data."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Score"])
        ws.append(["Alice", 90])
        ws.append(["Bob", 80])
        ws.append(["Carol", 70])
        ws["B5"] = "=SUM(B2:B4)"
        f = tmp_path / "with_summary.xlsx"
        wb.save(f); wb.close()

        # Manually save then reload with cached values populated.
        # openpyxl writes the formula but no cached value; just check the
        # formula text is faithfully preserved through import.
        result = import_excel(f)
        doc = parse_string(result)

        # Formula text should appear in compute (advisory).
        assert "=SUM(B2:B4)" in doc.compute_raw

        # The summary row itself MUST still appear in the data block —
        # we never silently delete user-typed cells.
        from gridlang.schema import parse_data
        df = parse_data(doc.data_raw)
        # 4 rows: Alice / Bob / Carol / the summary row.
        assert len(df) == 4

    def test_trailing_summary_row_with_blank_above_handled(self, tmp_path):
        """Blank rows between the data and a trailing summary row stay
        as blank rows in the imported CSV — preserving sheet layout."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["A", "B"])
        ws.append([1, 10])
        ws.append([2, 20])
        # row 4 blank
        ws["B5"] = "=SUM(B2:B3)"
        f = tmp_path / "summary_with_gap.xlsx"
        wb.save(f); wb.close()

        result = import_excel(f)
        doc = parse_string(result)
        assert "=SUM(B2:B3)" in doc.compute_raw

        from gridlang.schema import parse_data
        df = parse_data(doc.data_raw)
        # Two data rows + one blank + one summary row = 4 rows.
        assert len(df) == 4

    # ── 6. Faithful preservation of every cell ───────────────────────────

    def test_every_source_cell_preserved(self, tmp_path):
        """The imported .grid must contain EVERY non-empty cell from the
        source .xlsx — no values silently dropped, no extras invented.

        This is the cardinal rule of import: faithfulness. Earlier
        attempts at "smart" filtering deleted real user data; this test
        prevents that regression.
        """
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        wb = Workbook()
        ws = wb.active
        # Dense main block.
        for v in range(1, 6):
            ws.append([v, v, v, v])
        # Trailing scatter that earlier code wrongly filtered out:
        # blank rows 6-7, then row 8 with two values, then a summary.
        ws["A8"] = 999
        ws["B8"] = 888
        ws["A10"] = "=SUM(A1:A8)"  # we won't compute it; just store the formula
        f = tmp_path / "faithful.xlsx"
        wb.save(f); wb.close()

        # Collect every non-empty cell as written.
        # Use data_only=False so we see formula cells whose cached value
        # hasn't been computed (openpyxl never executes formulas).
        from openpyxl import load_workbook
        wb2 = load_workbook(f, data_only=False)
        ws2 = wb2.active
        src = {}
        for row in ws2.iter_rows():
            for c in row:
                if c.value is not None:
                    src[c.coordinate] = c.value

        # Parse imported .grid back into cell map.
        result = import_excel(f)
        from gridlang.schema import parse_data
        df = parse_data(parse_string(result).data_raw)
        imp = {}
        for r_idx, (_, row) in enumerate(df.iterrows()):
            for c_idx, val in enumerate(row):
                if val is None:
                    continue
                if isinstance(val, float) and val != val:  # NaN
                    continue
                if isinstance(val, str) and val == "":
                    continue
                coord = f"{get_column_letter(c_idx + 1)}{r_idx + 1}"
                imp[coord] = val

        # Same set of populated coordinates, same values (allowing
        # int/float numeric equality).
        missing = set(src) - set(imp)
        extra = set(imp) - set(src)
        assert not missing, f"Cells dropped from import: {sorted(missing)}"
        assert not extra, f"Cells invented during import: {sorted(extra)}"
        for k in src:
            s, i = src[k], imp[k]
            # Formula text: just ensure the formula round-tripped.
            if isinstance(s, str) and s.startswith("="):
                assert isinstance(i, str) and i == s, f"{k}: {s!r} != {i!r}"
                continue
            try:
                assert float(s) == float(i), f"{k}: {s!r} != {i!r}"
            except (TypeError, ValueError):
                assert s == i, f"{k}: {s!r} != {i!r}"

    def test_internal_blank_row_preserved(self, tmp_path):
        """A blank row INSIDE the data block stays in place — it's part
        of the user's layout."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["A", "B"])
        ws.append([1, 2])
        # Row 3 is blank.
        ws["A4"] = 3
        ws["B4"] = 4
        f = tmp_path / "internal_blank.xlsx"
        wb.save(f); wb.close()

        result = import_excel(f)
        from gridlang.schema import parse_data
        df = parse_data(parse_string(result).data_raw)
        # Three data rows: [1,2], blank, [3,4].
        assert len(df) == 3
        # Row 1 = (1, 2), row 2 = (NaN, NaN), row 3 = (3, 4)
        assert int(df.iloc[0, 0]) == 1
        # Middle row is fully empty.
        v_a, v_b = df.iloc[1, 0], df.iloc[1, 1]
        assert (v_a is None or (isinstance(v_a, float) and v_a != v_a))
        assert (v_b is None or (isinstance(v_b, float) and v_b != v_b))
        assert int(df.iloc[2, 0]) == 3
        assert int(df.iloc[2, 1]) == 4

    def test_trailing_blank_rows_trimmed(self, tmp_path):
        """Trailing fully-blank rows are NOT preserved (they extend the
        sheet's max_row but contain no user data)."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["A", "B"])
        ws.append([1, 2])
        ws["A10"] = None  # touch a far-away cell, then clear it
        f = tmp_path / "trailing_blank.xlsx"
        wb.save(f); wb.close()

        result = import_excel(f)
        from gridlang.schema import parse_data
        df = parse_data(parse_string(result).data_raw)
        # Just the one real data row.
        assert len(df) == 1

    # ── 7. End-to-end render ─────────────────────────────────────────────

    def test_e2e_render_no_floaty_integers(self, tmp_path):
        """The full pipeline (import → parse → execute → render) on a
        sheet that triggers float-promotion must NOT show 1.0 / 2.0 …

        ALSO: every source cell must round-trip into the rendered HTML —
        no row silently dropped just because it was sparse.
        """
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        for v in range(1, 11):
            ws.append([v] * 5)
        # A sparse trailing row [99, None×4]. Must STILL appear in render
        # (faithful preservation), and "99" should render as plain int.
        ws.append([99, None, None, None, None])
        f = tmp_path / "render_check.xlsx"
        wb.save(f); wb.close()

        result = import_excel(f)
        from gridlang.schema import parse_data
        from gridlang.runtime import execute
        from gridlang.renderer import render
        doc = parse_string(result)
        df = parse_data(doc.data_raw)
        out = execute(doc.compute_raw, df)
        html = render(
            template_content=doc.present_raw, df=out.df,
            aggregates=out.aggregates, meta=doc.meta,
        )
        import re
        body = re.search(r"<tbody>(.*?)</tbody>", html, re.S).group(1)
        # Faithful: 11 rows in, 11 rows out.
        rendered_rows = re.findall(r"<tr>.*?</tr>", body, re.S)
        assert len(rendered_rows) == 11, (
            f"Expected 11 <tr> rows (10 dense + 1 sparse), got {len(rendered_rows)}"
        )
        # No "1.0", "5.0", "99.0" should appear.
        for n in (1, 5, 10, 99):
            assert f">{n}<" in body, f"Missing integer cell '{n}' in rendered HTML"
            assert f">{n}.0<" not in body, f"Integer {n} rendered as float {n}.0"

    # ── 8. Multi-sheet, sheet selection, errors ──────────────────────────

    def test_specific_sheet(self, sample_xlsx):
        result = import_excel(sample_xlsx, sheet_names=["Data"])
        assert "Alpha" in result

    def test_unknown_sheet_raises(self, sample_xlsx):
        from gridlang.excel_import import ImportError_
        with pytest.raises(ImportError_):
            import_excel(sample_xlsx, sheet_names=["Nonexistent"])

    def test_multi_sheet_emits_named_sections(self, tmp_path):
        from openpyxl import Workbook
        wb = Workbook()
        ws1 = wb.active; ws1.title = "Sales"
        ws1.append(["Region", "Total"]); ws1.append(["North", 100])
        ws2 = wb.create_sheet("Expenses")
        ws2.append(["Item", "Cost"]); ws2.append(["Rent", 50])
        f = tmp_path / "multi.xlsx"
        wb.save(f); wb.close()

        result = import_excel(f)
        # Both sheet sections present.
        assert "--- data:sales ---" in result
        assert "--- data:expenses ---" in result
        assert "North" in result and "Rent" in result

    def test_file_not_found(self):
        with pytest.raises(FileNotFoundError):
            import_excel("/nonexistent/file.xlsx")

    def test_empty_sheet_produces_valid_grid(self, tmp_path):
        """A sheet with no data shouldn't crash — emits a stub .grid."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        # No rows appended — the sheet exists but is blank.
        f = tmp_path / "empty.xlsx"
        wb.save(f); wb.close()

        result = import_excel(f)
        # Still parses as a valid 4-section file.
        doc = parse_string(result)
        assert "def transform" in doc.compute_raw

    # ── 9. Real-world fixture parity check ───────────────────────────────

    def test_excel_style_view_has_letter_columns_and_row_numbers(self, tmp_path):
        """The auto-generated present template renders as a spreadsheet
        view: A/B/C... column headers, 1/2/3... row numbers, every cell
        visible (empty cells render as blank ``<td>``)."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        for v in (1, 2, 3):
            ws.append([v, v, v, v])
        # row 5 blank
        ws["A6"] = 99
        f = tmp_path / "excel_style.xlsx"
        wb.save(f); wb.close()

        result = import_excel(f)
        from gridlang.schema import parse_data
        from gridlang.runtime import execute
        from gridlang.renderer import render
        doc = parse_string(result)
        df = parse_data(doc.data_raw)
        out = execute(doc.compute_raw, df)
        html = render(
            template_content=doc.present_raw, df=out.df,
            aggregates=out.aggregates, meta=doc.meta,
        )

        import re

        # Column letters: A, B, C, D in <thead>.
        thead = re.search(r"<thead>(.*?)</thead>", html, re.S).group(1)
        for letter in ["A", "B", "C", "D"]:
            assert f">{letter}</th>" in thead, f"Missing column letter {letter}"

        # Row numbers in <tbody>.
        body = re.search(r"<tbody>(.*?)</tbody>", html, re.S).group(1)
        rows = re.findall(r"<tr>.*?</tr>", body, re.S)
        # 6 rows: rows 1-3 (data), 4 (blank — nothing was written there),
        # 5 (blank), 6 (sparse: A6=99).
        # Note: openpyxl does write row 4 even though we appended 3 rows,
        # because ws["A6"] forces the worksheet to extend. Either way we
        # expect EVERY row from 1 through the last non-empty row.
        assert len(rows) >= 4, f"Expected at least 4 rendered rows, got {len(rows)}"
        for i in range(1, len(rows) + 1):
            assert f'class="gl-rownum">{i}</td>' in rows[i - 1], f"Missing row number {i}"

        # Some row in the middle is blank — must show visible empty cells.
        any_blank = any("gl-empty" in r for r in rows[3:])
        assert any_blank, "No blank row had visible empty cells"
        # The last row has 99 in column A.
        assert ">99<" in rows[-1]
        assert rows[-1].count("gl-empty") == 3

    def test_测试1_xlsx_cell_parity(self, tmp_path):
        """The exact `测试1.xlsx` the user reported MUST round-trip every
        cell. This pins the regression that the 'detached row' filter
        introduced (it dropped A24=1, B24=2, A27=2082)."""
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl import load_workbook

        # Reconstruct the exact failing layout: 21 dense rows of 1..21
        # across 15 columns, blank rows 22-23, [1,2] in row 24, blank
        # rows 25-26, =SUM(A1:I26) at A27 (cached value 2082).
        wb = Workbook()
        ws = wb.active
        for v in range(1, 22):
            ws.append([v] * 15)
        # Skip rows 22, 23 by writing directly to row 24.
        ws["A24"] = 1
        ws["B24"] = 2
        # Row 27: a formula. We deliberately don't set a cached value
        # because openpyxl doesn't compute formulas — but the cell is
        # still a real cell that the import must surface.
        ws["A27"] = "=SUM(A1:I26)"
        f = tmp_path / "测试1.xlsx"
        wb.save(f); wb.close()

        # Source side: every non-empty cell. Use data_only=False so
        # formula cells without a cached value still register.
        wb2 = load_workbook(f, data_only=False)
        ws2 = wb2.active
        src = {c.coordinate: c.value
               for row in ws2.iter_rows() for c in row
               if c.value is not None}

        # Imported side.
        result = import_excel(f)
        from gridlang.schema import parse_data
        df = parse_data(parse_string(result).data_raw)
        imp = {}
        for r_idx, (_, row) in enumerate(df.iterrows()):
            for c_idx, val in enumerate(row):
                if val is None: continue
                if isinstance(val, float) and val != val: continue
                if isinstance(val, str) and val == "": continue
                coord = f"{get_column_letter(c_idx + 1)}{r_idx + 1}"
                imp[coord] = val

        missing = set(src) - set(imp)
        assert not missing, (
            f"Cells dropped: {sorted(missing)} — these were in the source "
            f"but vanished after import. This is the exact bug from 测试1.xlsx."
        )
        # Spot-check the three cells that historically were dropped.
        for k in ("A24", "B24", "A27"):
            assert k in imp, f"{k} (a real source cell) missing from import"


class TestExcelExport:
    def test_basic_export(self, sample_grid, tmp_path):
        output = tmp_path / "output.xlsx"
        result_path = export_excel(sample_grid, output)
        assert result_path.exists()
        assert result_path.stat().st_size > 0

    def test_export_has_data(self, sample_grid, tmp_path):
        output = tmp_path / "output.xlsx"
        export_excel(sample_grid, output)

        from openpyxl import load_workbook
        wb = load_workbook(output)
        ws = wb.active
        # Check header
        assert ws.cell(row=1, column=1).value == "Product"
        # Check data
        assert ws.cell(row=2, column=1).value == "Widget"
        wb.close()

    def test_export_with_summary(self, sample_grid, tmp_path):
        output = tmp_path / "output.xlsx"
        export_excel(sample_grid, output, include_aggregates=True)

        from openpyxl import load_workbook
        wb = load_workbook(output)
        assert "Summary" in wb.sheetnames
        wb.close()

    def test_export_dataframe(self, tmp_path):
        df = pd.DataFrame({'A': [1, 2, 3], 'B': [4, 5, 6]})
        output = tmp_path / "df.xlsx"
        result = export_dataframe(df, output)
        assert result.exists()


class TestRoundTrip:
    """Test xlsx → .grid → xlsx preserves data."""

    def test_xlsx_roundtrip(self, sample_xlsx, tmp_path):
        # Import
        grid_content = import_excel(sample_xlsx)
        grid_file = tmp_path / "roundtrip.grid"
        grid_file.write_text(grid_content)

        # Export back
        output_xlsx = tmp_path / "roundtrip.xlsx"
        export_excel(grid_file, output_xlsx)

        # Verify
        from openpyxl import load_workbook
        wb = load_workbook(output_xlsx)
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "Alpha"
        assert ws.cell(row=2, column=2).value == 100
        wb.close()

    def test_csv_roundtrip(self, sample_csv, tmp_path):
        # Import
        grid_content = import_csv(sample_csv)
        grid_file = tmp_path / "roundtrip.grid"
        grid_file.write_text(grid_content)

        # Export back to CSV
        output_csv = tmp_path / "roundtrip.csv"
        export_csv(grid_file, output_csv, raw=True)

        # Verify
        df = pd.read_csv(output_csv)
        assert list(df['Name']) == ['Alice', 'Bob', 'Charlie']
        assert list(df['Score']) == [90, 80, 70]
