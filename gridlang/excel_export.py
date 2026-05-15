"""
GridLang Excel Export — Convert .grid files to .xlsx format.

Handles:
- DataFrame data → Excel cells
- Conditional formats → Excel conditional formatting
- Number formats
- Header styling
- Multi-sheet support
- Aggregates → summary sheet
"""

from __future__ import annotations

from pathlib import Path
from typing import Optional, Any

import pandas as pd
import numpy as np

try:
    import xlsxwriter
    HAS_XLSXWRITER = True
except ImportError:
    HAS_XLSXWRITER = False

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.utils import get_column_letter

from gridlang.parser import parse_file, GridDocument
from gridlang.schema import parse_data
from gridlang.runtime import execute, ExecutionResult, ConditionalFormat


class ExportError(Exception):
    """Raised when Excel export fails."""
    pass


def export_excel(
    grid_path: str | Path,
    output_path: str | Path,
    include_aggregates: bool = True,
    include_formatting: bool = True,
    merge_cells: Optional[list[dict]] = None,
    engine: str = 'openpyxl',
) -> Path:
    """
    Export a .grid file to Excel format.

    Pipeline: parse .grid → execute compute → write .xlsx

    Args:
        grid_path: Path to .grid file.
        output_path: Output .xlsx path.
        include_aggregates: Add a summary sheet with aggregates.
        include_formatting: Apply conditional formats and styling.
        engine: 'openpyxl' (default) or 'xlsxwriter'.

    Returns:
        Path to the generated .xlsx file.
    """
    # Parse and execute
    doc = parse_file(grid_path)

    # Parse all sheets
    if doc.is_multi_sheet:
        sheets_data = {}
        for name, raw in doc.sheets_raw.items():
            sheets_data[name] = parse_data(raw)
    else:
        sheets_data = {'default': parse_data(doc.data_raw)}

    primary_df = list(sheets_data.values())[0]

    # Execute compute
    result = execute(doc.compute_raw, primary_df, sheets=sheets_data)

    # Write to Excel
    out_path = Path(output_path)

    if engine == 'xlsxwriter' and HAS_XLSXWRITER:
        _export_xlsxwriter(result, doc, out_path, include_aggregates, include_formatting)
    else:
        _export_openpyxl(result, doc, out_path, include_aggregates, include_formatting)

    return out_path


def export_dataframe(
    df: pd.DataFrame,
    output_path: str | Path,
    sheet_name: str = "Sheet1",
    title: str = "",
) -> Path:
    """Simple export: DataFrame → .xlsx with basic formatting."""
    out_path = Path(output_path)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    _write_dataframe_to_sheet(ws, df, include_formatting=True, title=title)

    wb.save(out_path)
    wb.close()
    return out_path


# =============================================================================
# OpenPyXL Export
# =============================================================================

def _export_openpyxl(
    result: ExecutionResult,
    doc: GridDocument,
    output_path: Path,
    include_aggregates: bool,
    include_formatting: bool,
):
    """Export using openpyxl engine."""
    wb = Workbook()

    # Write data sheets
    sheets_to_write = result.sheets if result.is_multi_sheet else {'Data': result.df}
    first_sheet = True

    for sheet_name, df in sheets_to_write.items():
        if first_sheet:
            ws = wb.active
            ws.title = sheet_name
            first_sheet = False
        else:
            ws = wb.create_sheet(title=sheet_name)

        _write_dataframe_to_sheet(ws, df, include_formatting)

        # Apply conditional formats
        if include_formatting and result.conditional_formats:
            _apply_conditional_formats_openpyxl(ws, df, result.conditional_formats)

    # Charts sheet — auto-generate from numeric data
    primary_df = result.df
    numeric_cols = primary_df.select_dtypes(include=[np.number]).columns.tolist()
    string_cols = primary_df.select_dtypes(include=['object']).columns.tolist()

    if numeric_cols and string_cols and include_formatting:
        _create_charts_sheet_openpyxl(wb, primary_df, numeric_cols, string_cols)

    # Aggregates sheet
    if include_aggregates and result.aggregates:
        ws_agg = wb.create_sheet(title="Summary")
        _write_aggregates_sheet(ws_agg, result.aggregates, doc.name)

    wb.save(output_path)
    wb.close()


def _write_dataframe_to_sheet(ws, df: pd.DataFrame, include_formatting: bool = True, title: str = ""):
    """Write DataFrame to worksheet with formatting."""
    start_row = 1

    # Optional title
    if title:
        ws.cell(row=1, column=1, value=title)
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        start_row = 3

    # Header row
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=str(col_name))
        if include_formatting:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

    # Data rows
    for row_idx, (_, row) in enumerate(df.iterrows(), start_row + 1):
        for col_idx, col_name in enumerate(df.columns, 1):
            value = row[col_name]

            # Handle NaN/None
            if pd.isna(value):
                ws.cell(row=row_idx, column=col_idx, value=None)
                continue

            # Convert numpy types to native Python
            if isinstance(value, (np.integer,)):
                value = int(value)
            elif isinstance(value, (np.floating,)):
                value = float(value)
            elif isinstance(value, (np.bool_,)):
                value = bool(value)
            elif isinstance(value, pd.Timestamp):
                value = value.to_pydatetime()

            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # Auto-detect number format
            if include_formatting and isinstance(value, (int, float)):
                if abs(value) >= 1000:
                    cell.number_format = '#,##0'
                elif isinstance(value, float) and abs(value) < 1:
                    cell.number_format = '0.0%'

    # Auto-adjust column widths
    for col_idx, col_name in enumerate(df.columns, 1):
        max_length = len(str(col_name))
        for row_idx in range(start_row + 1, min(start_row + 50, ws.max_row + 1)):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 30)

    # Freeze top row (header)
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

    # Auto-filter
    if df.shape[0] > 0:
        ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(df.columns))}{start_row + len(df)}"


def _apply_conditional_formats_openpyxl(ws, df: pd.DataFrame, rules: list[ConditionalFormat]):
    """Apply conditional formatting rules to worksheet."""
    for rule in rules:
        if rule.column not in df.columns:
            continue

        col_idx = list(df.columns).index(rule.column) + 1
        col_letter = get_column_letter(col_idx)
        cell_range = f"{col_letter}2:{col_letter}{len(df) + 1}"

        if rule.rule == 'greater_than' and rule.value is not None:
            ws.conditional_formatting.add(
                cell_range,
                CellIsRule(
                    operator='greaterThan',
                    formula=[str(rule.value)],
                    fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
                    font=Font(color="006100"),
                )
            )
        elif rule.rule == 'less_than' and rule.value is not None:
            ws.conditional_formatting.add(
                cell_range,
                CellIsRule(
                    operator='lessThan',
                    formula=[str(rule.value)],
                    fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
                    font=Font(color="9C0006"),
                )
            )
        elif rule.rule == 'color_scale':
            ws.conditional_formatting.add(
                cell_range,
                ColorScaleRule(
                    start_type='min', start_color=rule.min_color.lstrip('#'),
                    end_type='max', end_color=rule.max_color.lstrip('#'),
                )
            )
        elif rule.rule == 'data_bar':
            ws.conditional_formatting.add(
                cell_range,
                DataBarRule(
                    start_type='min', end_type='max',
                    color="3B82F6",
                )
            )


def _write_aggregates_sheet(ws, aggregates: dict, doc_name: str):
    """Write aggregates to a summary sheet."""
    # Title
    ws.cell(row=1, column=1, value=f"Summary: {doc_name}")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)

    # Headers
    ws.cell(row=3, column=1, value="Metric")
    ws.cell(row=3, column=2, value="Value")
    ws.cell(row=3, column=1).font = Font(bold=True)
    ws.cell(row=3, column=2).font = Font(bold=True)

    # Data
    for i, (key, value) in enumerate(aggregates.items(), 4):
        label = key.replace('_', ' ').title()
        ws.cell(row=i, column=1, value=label)

        # Convert value
        if isinstance(value, (np.integer,)):
            value = int(value)
        elif isinstance(value, (np.floating,)):
            value = float(value)
        elif isinstance(value, (list, pd.Series)):
            value = str(value)  # Skip list values for summary

        ws.cell(row=i, column=2, value=value)

        # Format numbers
        if isinstance(value, float):
            ws.cell(row=i, column=2).number_format = '#,##0.00'
        elif isinstance(value, int) and abs(value) >= 1000:
            ws.cell(row=i, column=2).number_format = '#,##0'

    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20


# =============================================================================
# XlsxWriter Export (higher quality charts/formatting)
# =============================================================================

def _export_xlsxwriter(
    result: ExecutionResult,
    doc: GridDocument,
    output_path: Path,
    include_aggregates: bool,
    include_formatting: bool,
):
    """Export using xlsxwriter engine (better for charts)."""
    workbook = xlsxwriter.Workbook(str(output_path))

    # Formats
    header_format = workbook.add_format({
        'bold': True, 'font_color': 'white', 'bg_color': '#1E40AF',
        'border': 1, 'align': 'center', 'valign': 'vcenter',
    })
    number_format = workbook.add_format({'num_format': '#,##0'})
    pct_format = workbook.add_format({'num_format': '0.0%'})
    currency_format = workbook.add_format({'num_format': '$#,##0'})

    # Write data sheets
    sheets_to_write = result.sheets if result.is_multi_sheet else {'Data': result.df}

    for sheet_name, df in sheets_to_write.items():
        ws = workbook.add_worksheet(sheet_name[:31])  # Excel 31-char limit

        # Headers
        for col_idx, col_name in enumerate(df.columns):
            ws.write(0, col_idx, str(col_name), header_format)

        # Data
        for row_idx, (_, row) in enumerate(df.iterrows(), 1):
            for col_idx, col_name in enumerate(df.columns):
                value = row[col_name]
                if pd.isna(value):
                    ws.write_blank(row_idx, col_idx, None)
                elif isinstance(value, (np.integer,)):
                    ws.write_number(row_idx, col_idx, int(value), number_format)
                elif isinstance(value, (np.floating, float)):
                    ws.write_number(row_idx, col_idx, float(value), number_format)
                elif isinstance(value, bool):
                    ws.write_boolean(row_idx, col_idx, value)
                else:
                    ws.write_string(row_idx, col_idx, str(value))

        # Auto-fit (approximate)
        for col_idx, col_name in enumerate(df.columns):
            width = max(len(str(col_name)), 8)
            ws.set_column(col_idx, col_idx, width + 2)

        # Freeze header
        ws.freeze_panes(1, 0)

        # Auto-filter
        ws.autofilter(0, 0, len(df), len(df.columns) - 1)

        # Conditional formats
        if include_formatting and result.conditional_formats:
            _apply_conditional_formats_xlsxwriter(ws, workbook, df, result.conditional_formats)

    # Summary sheet
    if include_aggregates and result.aggregates:
        ws_sum = workbook.add_worksheet("Summary")
        title_format = workbook.add_format({'bold': True, 'font_size': 14})
        ws_sum.write(0, 0, f"Summary: {doc.name}", title_format)
        ws_sum.write(2, 0, "Metric", header_format)
        ws_sum.write(2, 1, "Value", header_format)

        for i, (key, value) in enumerate(result.aggregates.items(), 3):
            label = key.replace('_', ' ').title()
            ws_sum.write(i, 0, label)
            if isinstance(value, (np.integer,)):
                ws_sum.write_number(i, 1, int(value), number_format)
            elif isinstance(value, (np.floating, float)):
                ws_sum.write_number(i, 1, float(value), number_format)
            else:
                ws_sum.write(i, 1, str(value))

        ws_sum.set_column(0, 0, 25)
        ws_sum.set_column(1, 1, 20)

    workbook.close()


def _apply_conditional_formats_xlsxwriter(ws, workbook, df: pd.DataFrame, rules: list[ConditionalFormat]):
    """Apply conditional formatting using xlsxwriter."""
    for rule in rules:
        if rule.column not in df.columns:
            continue

        col_idx = list(df.columns).index(rule.column)
        cell_range = f"{get_column_letter(col_idx + 1)}2:{get_column_letter(col_idx + 1)}{len(df) + 1}"

        if rule.rule == 'greater_than' and rule.value is not None:
            ws.conditional_format(1, col_idx, len(df), col_idx, {
                'type': 'cell', 'criteria': '>', 'value': rule.value,
                'format': workbook.add_format({'bg_color': '#C6EFCE', 'font_color': '#006100'}),
            })
        elif rule.rule == 'less_than' and rule.value is not None:
            ws.conditional_format(1, col_idx, len(df), col_idx, {
                'type': 'cell', 'criteria': '<', 'value': rule.value,
                'format': workbook.add_format({'bg_color': '#FFC7CE', 'font_color': '#9C0006'}),
            })
        elif rule.rule == 'color_scale':
            ws.conditional_format(1, col_idx, len(df), col_idx, {
                'type': '2_color_scale',
                'min_color': rule.min_color, 'max_color': rule.max_color,
            })
        elif rule.rule == 'data_bar':
            ws.conditional_format(1, col_idx, len(df), col_idx, {
                'type': 'data_bar', 'bar_color': '#3B82F6',
            })


# =============================================================================
# Native Excel Chart Generation (openpyxl)
# =============================================================================

def _create_charts_sheet_openpyxl(wb, df: pd.DataFrame, numeric_cols: list, string_cols: list):
    """Create a Charts sheet with native Excel charts based on the data."""
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference

    ws_charts = wb.create_sheet(title="Charts")

    # Write chart data to hidden area of the Charts sheet
    # Headers
    label_col = string_cols[0]  # Use first string column as labels
    chart_cols = numeric_cols[:4]  # Limit to 4 numeric columns for readability

    # Write data block for charts (starting at A1)
    ws_charts.cell(row=1, column=1, value=label_col)
    for j, col in enumerate(chart_cols, 2):
        ws_charts.cell(row=1, column=j, value=col)

    for i, (_, row) in enumerate(df.iterrows(), 2):
        ws_charts.cell(row=i, column=1, value=str(row[label_col]))
        for j, col in enumerate(chart_cols, 2):
            val = row[col]
            if pd.notna(val):
                try:
                    ws_charts.cell(row=i, column=j, value=float(val))
                except (ValueError, TypeError):
                    ws_charts.cell(row=i, column=j, value=val)

    n_rows = len(df) + 1  # +1 for header
    n_data_cols = len(chart_cols) + 1  # +1 for label column

    # --- Bar Chart ---
    bar_chart = BarChart()
    bar_chart.type = "col"
    bar_chart.title = f"{chart_cols[0]} by {label_col}"
    bar_chart.y_axis.title = chart_cols[0]
    bar_chart.x_axis.title = label_col
    bar_chart.style = 10
    bar_chart.width = 18
    bar_chart.height = 12

    data_ref = Reference(ws_charts, min_col=2, min_row=1, max_row=n_rows, max_col=min(n_data_cols, 4))
    cats_ref = Reference(ws_charts, min_col=1, min_row=2, max_row=n_rows)
    bar_chart.add_data(data_ref, titles_from_data=True)
    bar_chart.set_categories(cats_ref)
    bar_chart.shape = 4

    ws_charts.add_chart(bar_chart, "A" + str(n_rows + 3))

    # --- Line Chart (if enough data points) ---
    if len(df) >= 3 and len(chart_cols) >= 2:
        line_chart = LineChart()
        line_chart.title = "Trend"
        line_chart.style = 10
        line_chart.width = 18
        line_chart.height = 12

        line_data = Reference(ws_charts, min_col=2, min_row=1, max_row=n_rows, max_col=min(n_data_cols, 4))
        line_cats = Reference(ws_charts, min_col=1, min_row=2, max_row=n_rows)
        line_chart.add_data(line_data, titles_from_data=True)
        line_chart.set_categories(line_cats)

        ws_charts.add_chart(line_chart, "K" + str(n_rows + 3))

    # --- Pie Chart (first numeric column) ---
    if len(df) <= 12:  # Pie charts work best with limited slices
        pie_chart = PieChart()
        pie_chart.title = f"{chart_cols[0]} Distribution"
        pie_chart.style = 26
        pie_chart.width = 14
        pie_chart.height = 12

        pie_data = Reference(ws_charts, min_col=2, min_row=1, max_row=n_rows)
        pie_cats = Reference(ws_charts, min_col=1, min_row=2, max_row=n_rows)
        pie_chart.add_data(pie_data, titles_from_data=True)
        pie_chart.set_categories(pie_cats)

        chart_row = n_rows + 20
        ws_charts.add_chart(pie_chart, "A" + str(chart_row))
