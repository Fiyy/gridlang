"""
GridLang Excel Import — Convert .xlsx files to .grid format.

Handles:
- Data extraction (all sheets → CSV)
- Formula detection and conversion to Python (best-effort)
- Format extraction (fonts, colors, conditional formats) → HTML/CSS
- Multi-sheet workbooks → multi data sections
"""

from __future__ import annotations

import re
import io
from pathlib import Path
from typing import Optional
from datetime import datetime

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
try:
    from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
except ImportError:
    CellIsRule = ColorScaleRule = DataBarRule = None


class ImportError_(Exception):
    """Raised when Excel import fails."""
    pass


def import_excel(
    xlsx_path: str | Path,
    sheet_names: Optional[list[str]] = None,
    include_formulas: bool = True,
    include_styles: bool = True,
) -> str:
    """
    Convert an Excel file to .grid format.

    Args:
        xlsx_path: Path to .xlsx file.
        sheet_names: Specific sheets to import (None = all).
        include_formulas: Attempt to convert formulas to Python.
        include_styles: Extract formatting to HTML/CSS.

    Returns:
        String content of the resulting .grid file.
    """
    path = Path(xlsx_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")

    # Load workbook with data values and formulas
    wb_data = load_workbook(path, data_only=True)
    wb_formulas = load_workbook(path, data_only=False) if include_formulas else None

    # Determine which sheets to process
    available_sheets = wb_data.sheetnames
    if sheet_names:
        for name in sheet_names:
            if name not in available_sheets:
                raise ImportError_(f"Sheet '{name}' not found. Available: {available_sheets}")
        target_sheets = sheet_names
    else:
        target_sheets = available_sheets

    # Build .grid file content
    parts = []

    # --- meta ---
    parts.append("--- meta ---")
    parts.append(f'name: "{path.stem}"')
    parts.append('engine: python')
    parts.append('version: "1.0"')
    parts.append(f'description: "Imported from {path.name}"')
    parts.append(f'imported_from: "{path.name}"')
    parts.append(f'import_date: "{datetime.now().strftime("%Y-%m-%d %H:%M")}"')
    if len(target_sheets) > 1:
        sheets_str = ', '.join(f'"{s}"' for s in target_sheets)
        parts.append(f'sheets: [{sheets_str}]')
    parts.append("")

    # --- data (per sheet) ---
    all_formulas = {}  # sheet_name → list of formula info
    all_styles = {}    # sheet_name → style info
    all_summaries = {}  # sheet_name → list of {coord, formula, value}
    all_extras = {}     # sheet_name → list of {coord, value} (detached cells)

    for sheet_name in target_sheets:
        ws_data = wb_data[sheet_name]
        ws_formula = wb_formulas[sheet_name] if wb_formulas else None

        # Extract data as CSV
        if len(target_sheets) == 1:
            parts.append("--- data ---")
        else:
            safe_name = _safe_sheet_name(sheet_name)
            parts.append(f"--- data:{safe_name} ---")

        csv_content, summary_rows, extra_cells = _extract_sheet_data(ws_data, ws_formula)
        parts.append(csv_content)
        parts.append("")

        if summary_rows:
            all_summaries[sheet_name] = summary_rows
        if extra_cells:
            all_extras[sheet_name] = extra_cells

        # Collect formulas
        if ws_formula and include_formulas:
            formulas = _extract_formulas(ws_formula)
            if formulas:
                all_formulas[sheet_name] = formulas

        # Collect styles
        if include_styles:
            styles = _extract_styles(ws_data)
            if styles:
                all_styles[sheet_name] = styles

    # --- compute ---
    parts.append("--- compute ---")
    compute_code = _generate_compute_section(
        all_formulas, target_sheets, all_summaries, all_extras
    )
    parts.append(compute_code)
    parts.append("")

    # --- present ---
    parts.append("--- present ---")
    present_code = _generate_present_section(all_styles, target_sheets)
    parts.append(present_code)

    wb_data.close()
    if wb_formulas:
        wb_formulas.close()

    return "\n".join(parts)


def import_excel_to_file(
    xlsx_path: str | Path,
    output_path: str | Path,
    **kwargs,
) -> Path:
    """Import Excel and write to .grid file."""
    content = import_excel(xlsx_path, **kwargs)
    out = Path(output_path)
    out.write_text(content, encoding='utf-8')
    return out


# =============================================================================
# Data Extraction
# =============================================================================

def _looks_like_header_row(row: tuple) -> bool:
    """
    Decide whether a row of cells looks like a column-header row.

    True iff the row has at least one cell AND every non-empty cell is a
    string that's either:
      - non-numeric (e.g. "Region", "Q1 Sales", "%-of-total")
      - or short and obviously a label (no purely-numeric strings either).

    A row of all-numeric values (`1, 1, 1, ...` or `2024, 2025, ...`) is
    NOT a header — it's data.

    A mixed row ("Region", 100, 200, ...) is NOT a header either, because
    real headers don't mix labels and numbers in one row.
    """
    non_empty = [c for c in row if c is not None and str(c).strip() != ""]
    if not non_empty:
        return False  # empty rows are never headers

    for c in non_empty:
        # Numbers (int, float, bool) → not a header cell.
        if isinstance(c, (int, float, bool)) and not isinstance(c, str):
            return False
        # Strings that parse as numbers → not header cells.
        if isinstance(c, str):
            s = c.strip()
            try:
                float(s)
                return False  # purely numeric string
            except (ValueError, TypeError):
                pass
            if s.startswith("="):
                return False  # formula text — not a header
        else:
            # datetime, etc. — not a header
            return False
    return True


def _is_summary_row(row: tuple, formula_row: Optional[tuple] = None) -> bool:
    """
    Decide whether a row looks like a single-cell summary row at the bottom
    of a data block (e.g. just `=SUM(A1:I26)` in column A, rest empty).

    Heuristic: at most one non-empty cell, AND either that cell is a formula
    in `formula_row`, or the cached value in `row` is a number while every
    other cell is empty.
    """
    non_empty_idx = [i for i, c in enumerate(row) if c is not None and str(c).strip() != ""]
    if len(non_empty_idx) > 1:
        return False  # multiple values → looks like real data
    if not non_empty_idx:
        return False  # empty row, caller filters those separately
    # Exactly one value. If we have the formula-mode row, check it's a formula.
    if formula_row is not None:
        f = formula_row[non_empty_idx[0]]
        if isinstance(f, str) and f.startswith("="):
            return True
        # ArrayFormula objects also count as formulas.
        if f is not None and not isinstance(f, (int, float, bool, str)):
            return True
    return False


def _extract_sheet_data(ws, ws_formula=None) -> tuple[str, list[dict], list[dict]]:
    """Extract sheet data as a CSV string that **faithfully preserves**
    every cell the user typed.

    Design principle: never silently drop a value. The user's source of
    truth is the spreadsheet; the imported `.grid` must contain the same
    cells. Subsequent layers (compute / present) are advisory.

    Returns ``(csv, summary_rows, extra_cells)``. `summary_rows` is kept
    only as a *non-destructive* annotation: we still write the cached
    value of trailing single-cell formula rows into the CSV, but ALSO
    surface the formula text in compute as a hint. `extra_cells` is
    always returned as ``[]`` — kept in the signature for back-compat
    with callers; we no longer remove "detached" rows from data.
    """
    # ── 0. Snapshot every row exactly as openpyxl sees it.
    raw_rows = list(ws.iter_rows(values_only=True))
    formula_rows = (
        list(ws_formula.iter_rows(values_only=True))
        if ws_formula is not None
        else [None] * len(raw_rows)
    )

    if not raw_rows:
        return "", [], []

    # Determine the rectangle: trim trailing fully-blank rows but keep
    # blanks INSIDE the data block (they're meaningful in spreadsheets).
    # IMPORTANT: a row that's empty in data_only=True view but contains a
    # formula in data_only=False view is NOT blank — it has a real cell
    # whose cached value just hasn't been computed yet.
    def _row_is_blank(row, f_row):
        if row is not None and any(c is not None and str(c).strip() != "" for c in row):
            return False
        if f_row is not None and any(c is not None and str(c).strip() != "" for c in f_row):
            return False
        return True

    end = len(raw_rows)
    while end > 0:
        f_row = formula_rows[end - 1] if end - 1 < len(formula_rows) else None
        if _row_is_blank(raw_rows[end - 1], f_row):
            end -= 1
        else:
            break
    if end == 0:
        return "", [], []

    body_rows = raw_rows[:end]
    formula_body = formula_rows[:end] if formula_rows else [None] * end

    # ── 1. Note (don't drop) any trailing single-cell formula rows.
    # We emit the cached value into the CSV like every other cell; the
    # formula text still gets surfaced in the compute hints.
    summary_rows = []
    for i in range(len(body_rows)):
        cur = body_rows[i]
        f_row = formula_body[i] if i < len(formula_body) else None
        if _is_summary_row(cur, f_row):
            non_empty_idx = [j for j, c in enumerate(cur)
                             if c is not None and str(c).strip() != ""]
            if not non_empty_idx:
                continue
            j = non_empty_idx[0]
            coord = f"{get_column_letter(j + 1)}{i + 1}"
            f_val = f_row[j] if f_row is not None else None
            if hasattr(f_val, "text"):
                f_text = f_val.text
                if not (isinstance(f_text, str) and f_text.startswith("=")):
                    f_text = "=" + (f_text or "")
            elif isinstance(f_val, str) and f_val.startswith("="):
                f_text = f_val
            else:
                f_text = None
            summary_rows.append({
                "coord": coord,
                "formula": f_text,
                "value": cur[j],
            })

    # ── 2. Compute the column count from the maximum-width row.
    n_cols = max((len(r) for r in body_rows), default=0)
    if n_cols == 0:
        return "", summary_rows, []

    # Pad every row to n_cols so the CSV is rectangular.
    padded = [r + (None,) * (n_cols - len(r)) for r in body_rows]

    # ── 3. Detect whether row 1 is a textual header.
    has_header = _looks_like_header_row(padded[0])
    if has_header:
        header_row = padded[0]
        data_rows = padded[1:]
        headers = [_clean_column_name(str(c) if c is not None else "")
                   for c in header_row]
    else:
        headers = [f"col_{get_column_letter(i + 1)}" for i in range(n_cols)]
        data_rows = padded

    # ── 4. Per-column integer detection so whole numbers serialize as
    # `5` instead of `5.0`. Only fires if the WHOLE column is whole-
    # numbered (or empty).
    col_is_int = []
    for col in range(n_cols):
        all_int = True
        seen_any = False
        for row in data_rows:
            v = row[col] if col < len(row) else None
            if v is None or (isinstance(v, str) and v.strip() == ""):
                continue
            seen_any = True
            if isinstance(v, bool):
                all_int = False; break
            if isinstance(v, int):
                continue
            if isinstance(v, float):
                if v != v:    # NaN counts as empty
                    continue
                if v == int(v):
                    continue
                all_int = False; break
            all_int = False; break
        col_is_int.append(seen_any and all_int)

    # ── 5. Serialize. Every row is preserved.
    # If a cell is empty in the data-only view but has a formula in the
    # formula view, fall back to the formula text — that's still real
    # cell content (just an uncomputed result), not a blank.
    # We track formula-rows aligned to data_rows for this purpose.
    if has_header:
        formula_data_rows = formula_body[1:] if formula_body else [None] * len(data_rows)
    else:
        formula_data_rows = formula_body if formula_body else [None] * len(data_rows)
    # Pad formula rows to n_cols too.
    formula_data_rows = [
        (r + (None,) * (n_cols - len(r))) if r is not None
        else tuple([None] * n_cols)
        for r in formula_data_rows
    ]

    csv_lines = [",".join(headers)]
    for row_idx, row in enumerate(data_rows):
        f_row = formula_data_rows[row_idx] if row_idx < len(formula_data_rows) else None
        csv_row = []
        for col_idx, cell in enumerate(row):
            f_cell = f_row[col_idx] if f_row is not None and col_idx < len(f_row) else None
            # Fall back to formula text when the data view is empty but
            # there's a formula in this position.
            if (cell is None or (isinstance(cell, str) and cell == "")) and f_cell is not None:
                if hasattr(f_cell, "text") and isinstance(f_cell.text, str):
                    cell = f_cell.text if f_cell.text.startswith("=") else "=" + f_cell.text
                elif isinstance(f_cell, str) and f_cell.startswith("="):
                    cell = f_cell
                # else: leave cell as None

            if cell is None:
                csv_row.append("")
            elif isinstance(cell, datetime):
                csv_row.append(cell.strftime("%Y-%m-%d"))
            elif col_idx < len(col_is_int) and col_is_int[col_idx] and isinstance(cell, (int, float)):
                if isinstance(cell, float) and cell != cell:  # NaN
                    csv_row.append("")
                elif isinstance(cell, bool):
                    csv_row.append(str(cell))
                else:
                    csv_row.append(str(int(cell)))
            elif isinstance(cell, float) and not (cell != cell) and cell == int(cell):
                csv_row.append(str(int(cell)))
            else:
                val = str(cell)
                if "," in val or '"' in val or "\n" in val:
                    val = '"' + val.replace('"', '""') + '"'
                csv_row.append(val)
        csv_lines.append(",".join(csv_row))

    # extra_cells is intentionally always [] — we keep the parameter for
    # caller compatibility but no longer remove rows from the data block.
    return "\n".join(csv_lines), summary_rows, []


def _clean_column_name(name: str) -> str:
    """Convert column header to valid Python identifier."""
    name = name.strip().strip('"')
    # Replace spaces and special chars with underscore
    name = re.sub(r'[^\w]', '_', name)
    # Remove leading digits
    name = re.sub(r'^(\d)', r'_\1', name)
    # Remove multiple underscores
    name = re.sub(r'_+', '_', name).strip('_')
    return name or 'Column'


# =============================================================================
# Formula Extraction & Conversion
# =============================================================================

# Excel formula → Python conversion patterns
FORMULA_PATTERNS = [
    # Basic aggregations
    (r'=SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)', r"df['{col}'].sum()"),
    (r'=AVERAGE\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)', r"df['{col}'].mean()"),
    (r'=COUNT\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)', r"df['{col}'].count()"),
    (r'=MAX\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)', r"df['{col}'].max()"),
    (r'=MIN\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)', r"df['{col}'].min()"),
    # IF
    (r'=IF\(([^,]+),([^,]+),([^)]+)\)', r"IF({0}, {1}, {2})"),
    # VLOOKUP
    (r'=VLOOKUP\(', 'VLOOKUP('),
    # SUMIF
    (r'=SUMIF\(', 'SUMIF('),
    (r'=COUNTIF\(', 'COUNTIF('),
]


def _extract_formulas(ws) -> list[dict]:
    """Extract formulas from worksheet.

    Handles both plain `=SUM(...)` strings and openpyxl ArrayFormula
    objects (dynamic-array formulas). The latter would otherwise be
    silently dropped because their `.value` is not a string.
    """
    formulas = []
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if v is None:
                continue
            # Plain text formula.
            if isinstance(v, str) and v.startswith('='):
                formula_text = v
            # openpyxl ArrayFormula object — has a .text attribute.
            elif hasattr(v, 'text') and isinstance(getattr(v, 'text', None), str):
                formula_text = v.text
                if not formula_text.startswith('='):
                    formula_text = '=' + formula_text
            else:
                continue
            formulas.append({
                'cell': cell.coordinate,
                'column': get_column_letter(cell.column),
                'col_idx': cell.column - 1,
                'row': cell.row,
                'formula': formula_text,
            })
    return formulas


def _convert_formula_to_python(formula: str, headers: list[str] = None) -> str:
    """
    Best-effort conversion of Excel formula to Python expression.
    Returns Python code string, or a comment if conversion isn't possible.

    Supports:
    - Aggregations: SUM, AVERAGE, COUNT, MAX, MIN, STDEV, MEDIAN
    - Arithmetic: A1+B1, A1*2, A1/B1
    - IF: =IF(A1>0, "Yes", "No")
    - VLOOKUP: =VLOOKUP(A1, B:D, 2, FALSE)
    - SUMIF/COUNTIF: =SUMIF(A:A, ">10", B:B)
    - Text: CONCATENATE, LEFT, RIGHT, LEN, TRIM, UPPER, LOWER
    - Math: ROUND, ABS, MOD, POWER, CEILING, FLOOR
    - Date: YEAR, MONTH, DAY, TODAY, NOW
    - Nested formulas (basic level)
    """
    f = formula.strip()
    if not f.startswith('='):
        return f"# Not a formula: {formula}"

    f = f[1:]  # Remove leading '='

    def col_ref(letter):
        """Convert column letter to column name."""
        if headers:
            idx = ord(letter.upper()) - ord('A')
            if idx < len(headers):
                return headers[idx]
        return f"col_{letter}"

    # --- Simple column range aggregations: SUM(B2:B100) ---
    range_match = re.match(r'(\w+)\(([A-Z]+)\d+:([A-Z]+)\d+\)', f)
    if range_match:
        func = range_match.group(1).upper()
        col_letter = range_match.group(2)
        func_map = {
            'SUM': 'sum()', 'AVERAGE': 'mean()', 'COUNT': 'count()',
            'MAX': 'max()', 'MIN': 'min()', 'STDEV': 'std()',
            'MEDIAN': 'median()', 'VAR': 'var()',
            'COUNTA': 'count()', 'COUNTBLANK': 'isna().sum()',
        }
        if func in func_map:
            cn = col_ref(col_letter)
            return f"df['{cn}'].{func_map[func]}"

    # --- IF(condition, true_val, false_val) ---
    if_match = re.match(r'IF\((.+)\)', f, re.IGNORECASE)
    if if_match:
        inner = if_match.group(1)
        parts = _split_formula_args(inner)
        if len(parts) == 3:
            cond = _convert_condition(parts[0].strip(), headers)
            true_val = _convert_value(parts[1].strip(), headers)
            false_val = _convert_value(parts[2].strip(), headers)
            return f"IF({cond}, {true_val}, {false_val})"

    # --- VLOOKUP(value, table, col_idx, exact) ---
    vlookup_match = re.match(r'VLOOKUP\((.+)\)', f, re.IGNORECASE)
    if vlookup_match:
        parts = _split_formula_args(vlookup_match.group(1))
        if len(parts) >= 3:
            lookup_val = _convert_value(parts[0].strip(), headers)
            col_idx = parts[2].strip()
            return f"VLOOKUP({lookup_val}, df, {col_idx})"

    # --- SUMIF(range, criteria, [sum_range]) ---
    sumif_match = re.match(r'SUMIF\((.+)\)', f, re.IGNORECASE)
    if sumif_match:
        parts = _split_formula_args(sumif_match.group(1))
        if len(parts) >= 2:
            col = _convert_range_to_col(parts[0].strip(), headers)
            criteria = parts[1].strip()
            if len(parts) >= 3:
                sum_col = _convert_range_to_col(parts[2].strip(), headers)
                return f"SUMIF(df['{col}'], {criteria}, df['{sum_col}'])"
            return f"SUMIF(df['{col}'], {criteria})"

    # --- COUNTIF(range, criteria) ---
    countif_match = re.match(r'COUNTIF\((.+)\)', f, re.IGNORECASE)
    if countif_match:
        parts = _split_formula_args(countif_match.group(1))
        if len(parts) >= 2:
            col = _convert_range_to_col(parts[0].strip(), headers)
            criteria = parts[1].strip()
            return f"COUNTIF(df['{col}'], {criteria})"

    # --- CONCATENATE(...) ---
    concat_match = re.match(r'CONCATENATE\((.+)\)', f, re.IGNORECASE)
    if concat_match:
        parts = _split_formula_args(concat_match.group(1))
        converted = [_convert_value(p.strip(), headers) for p in parts]
        return f"CONCATENATE({', '.join(converted)})"

    # --- Text functions: LEFT, RIGHT, MID, LEN, TRIM, UPPER, LOWER ---
    text_funcs = {'LEFT': 'LEFT', 'RIGHT': 'RIGHT', 'MID': 'MID',
                  'LEN': 'LEN', 'TRIM': 'TRIM', 'UPPER': 'UPPER', 'LOWER': 'LOWER'}
    for excel_fn, py_fn in text_funcs.items():
        match = re.match(rf'{excel_fn}\((.+)\)', f, re.IGNORECASE)
        if match:
            args = _split_formula_args(match.group(1))
            converted_args = [_convert_value(a.strip(), headers) for a in args]
            return f"{py_fn}({', '.join(converted_args)})"

    # --- ROUND, ABS, MOD, POWER ---
    math_funcs = {'ROUND': 'ROUND', 'ROUNDUP': 'ROUNDUP', 'ROUNDDOWN': 'ROUNDDOWN',
                  'ABS': 'ABS', 'MOD': 'MOD', 'POWER': 'POWER',
                  'CEILING': 'CEILING', 'FLOOR': 'FLOOR'}
    for excel_fn, py_fn in math_funcs.items():
        match = re.match(rf'{excel_fn}\((.+)\)', f, re.IGNORECASE)
        if match:
            args = _split_formula_args(match.group(1))
            converted_args = [_convert_value(a.strip(), headers) for a in args]
            return f"{py_fn}({', '.join(converted_args)})"

    # --- Date functions: YEAR, MONTH, DAY ---
    date_funcs = {'YEAR': 'YEAR', 'MONTH': 'MONTH', 'DAY': 'DAY',
                  'TODAY': 'TODAY', 'NOW': 'NOW'}
    for excel_fn, py_fn in date_funcs.items():
        match = re.match(rf'{excel_fn}\((.+)?\)', f, re.IGNORECASE)
        if match:
            inner = match.group(1)
            if inner:
                arg = _convert_value(inner.strip(), headers)
                return f"{py_fn}({arg})"
            return f"{py_fn}()"

    # --- Simple arithmetic: A1+B1, A1*2, (A1-B1)/C1 ---
    arith_match = re.match(r'([A-Z]+)(\d+)\s*([+\-*/])\s*([A-Z]+)(\d+)', f)
    if arith_match:
        col1 = col_ref(arith_match.group(1))
        op = arith_match.group(3)
        col2 = col_ref(arith_match.group(4))
        return f"df['{col1}'] {op} df['{col2}']"

    # Column * constant: =A1*100, =B2/12
    arith_const = re.match(r'([A-Z]+)(\d+)\s*([+\-*/])\s*([\d.]+)', f)
    if arith_const:
        col1 = col_ref(arith_const.group(1))
        op = arith_const.group(3)
        num = arith_const.group(4)
        return f"df['{col1}'] {op} {num}"

    # Constant * column: =100*A1
    arith_const2 = re.match(r'([\d.]+)\s*([+\-*/])\s*([A-Z]+)(\d+)', f)
    if arith_const2:
        num = arith_const2.group(1)
        op = arith_const2.group(2)
        col1 = col_ref(arith_const2.group(3))
        return f"{num} {op} df['{col1}']"

    # --- Cannot convert — return as comment ---
    return f"# Excel: ={f}"


def _split_formula_args(s: str) -> list[str]:
    """Split formula arguments respecting nested parentheses and quotes."""
    parts = []
    depth = 0
    in_string = False
    current = []

    for ch in s:
        if ch == '"' and depth == 0:
            in_string = not in_string
            current.append(ch)
        elif ch == '(' and not in_string:
            depth += 1
            current.append(ch)
        elif ch == ')' and not in_string:
            depth -= 1
            current.append(ch)
        elif ch == ',' and depth == 0 and not in_string:
            parts.append(''.join(current))
            current = []
        else:
            current.append(ch)

    if current:
        parts.append(''.join(current))

    return parts


def _convert_condition(cond: str, headers: list[str] = None) -> str:
    """Convert an Excel condition to Python."""
    # Cell reference comparisons: A1>0, B2="Yes"
    match = re.match(r'([A-Z]+)(\d+)\s*(>=|<=|<>|>|<|=)\s*(.+)', cond)
    if match:
        col_letter = match.group(1)
        op = match.group(3)
        value = match.group(4).strip()

        col_name = col_letter
        if headers:
            idx = ord(col_letter.upper()) - ord('A')
            if idx < len(headers):
                col_name = headers[idx]

        # Convert operator
        op_map = {'=': '==', '<>': '!=', '>': '>', '<': '<', '>=': '>=', '<=': '<='}
        py_op = op_map.get(op, op)

        return f"df['{col_name}'] {py_op} {value}"

    return cond


def _convert_value(val: str, headers: list[str] = None) -> str:
    """Convert a value reference to Python."""
    # Cell reference: A1, B2
    match = re.match(r'^([A-Z]+)(\d+)$', val)
    if match:
        col_letter = match.group(1)
        if headers:
            idx = ord(col_letter.upper()) - ord('A')
            if idx < len(headers):
                return f"df['{headers[idx]}']"
        return f"df['col_{col_letter}']"

    # Already a string literal or number
    return val


def _convert_range_to_col(range_str: str, headers: list[str] = None) -> str:
    """Convert A:A or A1:A100 to column name."""
    match = re.match(r'([A-Z]+)(?::\1|\d*:[A-Z]+\d*)', range_str)
    if match:
        col_letter = match.group(1)
        if headers:
            idx = ord(col_letter.upper()) - ord('A')
            if idx < len(headers):
                return headers[idx]
        return f"col_{col_letter}"
    return range_str


def _generate_compute_section(
    all_formulas: dict,
    sheet_names: list[str],
    all_summaries: Optional[dict] = None,
    all_extras: Optional[dict] = None,
) -> str:
    """Generate compute section from extracted formulas."""
    lines = []
    lines.append("def transform(df):")

    has_formulas = any(formulas for formulas in all_formulas.values())
    has_summaries = bool(all_summaries) and any(s for s in (all_summaries or {}).values())
    has_extras = bool(all_extras) and any(e for e in (all_extras or {}).values())

    if has_formulas or has_summaries or has_extras:
        lines.append("    # Auto-converted from Excel formulas")
        lines.append("    # Review and adjust as needed")
        lines.append("")

        # Summaries (trailing single-cell formula rows) come first as a clearly
        # labelled block, since they were extracted out of the data table.
        if has_summaries:
            for sheet_name, summaries in all_summaries.items():
                if not summaries:
                    continue
                if len(all_summaries) > 1:
                    lines.append(f"    # --- Summary cells for sheet: {sheet_name} ---")
                else:
                    lines.append("    # --- Summary cells (trailing rows below data) ---")
                for s in summaries:
                    coord = s.get("coord", "?")
                    formula = s.get("formula") or "(array formula)"
                    value = s.get("value")
                    py_hint = _convert_formula_to_python(formula) if formula else ""
                    lines.append(f"    # Cell {coord}: {formula}  (cached value: {value})")
                    if py_hint and not py_hint.startswith("#"):
                        lines.append(f"    # → {py_hint}")
                lines.append("")

        # Detached "extra" cells — sparse rows that were below the main data block
        if has_extras:
            for sheet_name, extras in all_extras.items():
                if not extras:
                    continue
                if len(all_extras) > 1:
                    lines.append(f"    # --- Detached cells for sheet: {sheet_name} ---")
                else:
                    lines.append("    # --- Detached cells (separated from main data block by blank rows) ---")
                for e in extras:
                    lines.append(f"    # Cell {e['coord']}: {e['value']!r}")
                lines.append("")

        if has_formulas:
            for sheet_name, formulas in all_formulas.items():
                if len(all_formulas) > 1:
                    lines.append(f"    # --- Sheet: {sheet_name} ---")

                # Group formulas by type
                seen_patterns = set()
                for f_info in formulas:
                    python_code = _convert_formula_to_python(f_info['formula'])
                    # Avoid duplicates (same formula applied to many rows)
                    pattern_key = re.sub(r'\d+', 'N', f_info['formula'])
                    if pattern_key not in seen_patterns:
                        seen_patterns.add(pattern_key)
                        if python_code.startswith('#'):
                            lines.append(f"    {python_code}")
                        else:
                            lines.append(f"    # Cell {f_info['cell']}: {f_info['formula']}")
                            lines.append(f"    # → {python_code}")
            lines.append("")
    else:
        lines.append("    # No formulas detected — add your transformations here")
        lines.append("    pass")
        lines.append("")

    lines.append("    return df")

    return "\n".join(lines)


# =============================================================================
# Style Extraction
# =============================================================================

def _extract_styles(ws) -> dict:
    """Extract basic styling information from worksheet."""
    styles = {
        'has_bold_header': False,
        'has_colors': False,
        'number_formats': {},
        'conditional_formats': [],
        'column_widths': {},
        'merged_cells': [],
    }

    # Check header row styling
    first_row = list(ws.iter_rows(min_row=1, max_row=1))[0] if ws.max_row else []
    for cell in first_row:
        if cell.font and cell.font.bold:
            styles['has_bold_header'] = True
            break

    # Check for number formats
    for row in ws.iter_rows(min_row=2, max_row=min(10, ws.max_row or 1)):
        for cell in row:
            if cell.number_format and cell.number_format != 'General':
                col_letter = get_column_letter(cell.column)
                styles['number_formats'][col_letter] = cell.number_format

    # Extract conditional formatting rules
    try:
        for cf_rule in ws.conditional_formatting:
            for rule in cf_rule.rules:
                cf_info = {'range': str(cf_rule)}
                rule_type = type(rule).__name__
                if rule_type == 'CellIsRule':
                    cf_info['type'] = 'cell_is'
                    cf_info['operator'] = getattr(rule, 'operator', '')
                    cf_info['formula'] = rule.formula if hasattr(rule, 'formula') and rule.formula else []
                elif rule_type == 'ColorScaleRule':
                    cf_info['type'] = 'color_scale'
                elif rule_type == 'DataBarRule':
                    cf_info['type'] = 'data_bar'
                else:
                    cf_info['type'] = rule_type
                styles['conditional_formats'].append(cf_info)
    except Exception:
        pass  # Skip conditional format extraction on error

    # Column widths
    for col_dim in ws.column_dimensions.values():
        if col_dim.width:
            styles['column_widths'][col_dim.index] = col_dim.width

    # Merged cells
    for merged_range in ws.merged_cells.ranges:
        styles['merged_cells'].append({
            'range': str(merged_range),
            'min_row': merged_range.min_row,
            'max_row': merged_range.max_row,
            'min_col': merged_range.min_col,
            'max_col': merged_range.max_col,
            'colspan': merged_range.max_col - merged_range.min_col + 1,
            'rowspan': merged_range.max_row - merged_range.min_row + 1,
        })

    return styles


def _generate_present_section(all_styles: dict, sheet_names: list[str]) -> str:
    """Generate an Excel-style spreadsheet view as the present section.

    The view shows the FULL data block — every row, every column — with
    Excel-style column letters (A, B, C, ...) and row numbers as
    headers. Empty cells render as visible blanks. Integer values stay
    integer (no `1.0` promotion artifacts).
    """
    lines = []

    lines.append('<style>')
    lines.append('  .gl-sheet-wrap {')
    lines.append('    overflow: auto; border: 1px solid #c0c0c0;')
    lines.append('    background: #fff; max-height: 80vh;')
    lines.append('    font-family: "Segoe UI", "Helvetica Neue", Arial, sans-serif;')
    lines.append('  }')
    lines.append('  .gl-sheet { border-collapse: separate; border-spacing: 0; }')
    lines.append('  .gl-sheet th, .gl-sheet td {')
    lines.append('    border-right: 1px solid #d4d4d4; border-bottom: 1px solid #d4d4d4;')
    lines.append('    padding: 4px 8px; min-width: 64px; max-width: 220px;')
    lines.append('    font-size: 13px; line-height: 1.4; height: 22px;')
    lines.append('    white-space: nowrap; overflow: hidden; text-overflow: ellipsis;')
    lines.append('  }')
    lines.append('  /* Column-letter header row (A, B, C, ...) */')
    lines.append('  .gl-sheet thead th {')
    lines.append('    background: #f1f3f5; color: #333; font-weight: 600;')
    lines.append('    text-align: center; position: sticky; top: 0; z-index: 2;')
    lines.append('    border-top: 1px solid #d4d4d4;')
    lines.append('  }')
    lines.append('  /* Row-number cell on the left */')
    lines.append('  .gl-sheet th.gl-rownum, .gl-sheet td.gl-rownum {')
    lines.append('    background: #f1f3f5; color: #333; font-weight: 500;')
    lines.append('    text-align: center; position: sticky; left: 0;')
    lines.append('    min-width: 48px; max-width: 48px; z-index: 1;')
    lines.append('  }')
    lines.append('  .gl-sheet thead th.gl-corner {')
    lines.append('    z-index: 3; left: 0; min-width: 48px; max-width: 48px;')
    lines.append('  }')
    lines.append('  /* Data cells */')
    lines.append('  .gl-sheet tbody td.gl-cell {')
    lines.append('    background: #fff; color: #1f2937;')
    lines.append('    text-align: right;  /* numbers right-aligned by default, like Excel */')
    lines.append('    font-variant-numeric: tabular-nums;')
    lines.append('  }')
    lines.append('  .gl-sheet tbody td.gl-cell.gl-text { text-align: left; }')
    lines.append('  .gl-sheet tbody td.gl-cell.gl-empty { background: #fdfdfd; }')
    lines.append('  .gl-sheet tbody tr:hover td.gl-cell { background: #fff8c4; }')
    lines.append('  .gl-sheet tbody tr:hover td.gl-rownum { background: #e8edf2; }')
    lines.append('  .gl-meta { color: #6b7280; font-size: 12px; margin: 4px 0 8px; }')
    lines.append('</style>')
    lines.append('')
    lines.append('<h1>{{ meta.name }}</h1>')
    lines.append('<p class="gl-meta">Imported from {{ meta.imported_from }} '
                 '— {{ df.shape[0] }} rows × {{ df.shape[1] }} columns</p>')
    lines.append('')

    # Multi-sheet tabs
    if len(sheet_names) > 1:
        lines.append('<div class="sheet-tabs">')
        for i, name in enumerate(sheet_names):
            active = ' active' if i == 0 else ''
            lines.append(f'  <div class="sheet-tab{active}">{name}</div>')
        lines.append('</div>')
        lines.append('')

    # Excel-style sheet view: column-letter headers + row numbers, every
    # cell visible, empty cells render as blank.
    # Compute the column letters as a literal Jinja list so the template
    # doesn't need a `chr` filter. 702 letters covers A..ZZ — plenty.
    letters_list = [get_column_letter(i + 1) for i in range(702)]
    letters_literal = "[" + ", ".join(repr(l) for l in letters_list) + "]"

    lines.append('{% set _letters = ' + letters_literal + ' %}')
    lines.append('<div class="gl-sheet-wrap">')
    lines.append('<table class="gl-sheet">')
    lines.append('  <thead>')
    lines.append('    <tr>')
    lines.append('      <th class="gl-rownum gl-corner"></th>')
    lines.append('      {% for col in df.columns %}')
    lines.append('        <th title="{{ col }}">{{ _letters[loop.index0] }}</th>')
    lines.append('      {% endfor %}')
    lines.append('    </tr>')
    lines.append('  </thead>')
    lines.append('  <tbody>')
    lines.append('    {% for _, row in df.iterrows() %}')
    lines.append('    <tr>')
    lines.append('      <td class="gl-rownum">{{ loop.index }}</td>')
    lines.append('      {% for col in df.columns %}')
    lines.append('      {% set v = row[col] %}')
    # NaN/None → empty visible cell. Whole-number floats → int. Numbers
    # right-aligned, text left-aligned.
    lines.append('      {% if v is none or (v is number and v != v) %}'
                 '<td class="gl-cell gl-empty">&nbsp;</td>'
                 '{% elif v is number and (v | int) == v %}'
                 '<td class="gl-cell">{{ v | int }}</td>'
                 '{% elif v is number %}'
                 '<td class="gl-cell">{{ v }}</td>'
                 '{% else %}'
                 '<td class="gl-cell gl-text">{{ v }}</td>'
                 '{% endif %}')
    lines.append('      {% endfor %}')
    lines.append('    </tr>')
    lines.append('    {% endfor %}')
    lines.append('  </tbody>')
    lines.append('</table>')
    lines.append('</div>')

    return '\n'.join(lines)


def _safe_sheet_name(name: str) -> str:
    """Convert sheet name to safe section identifier."""
    safe = re.sub(r'[^\w]', '_', name.strip())
    safe = re.sub(r'_+', '_', safe).strip('_').lower()
    return safe or 'sheet'
